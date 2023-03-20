import seaborn as sns
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path
import matplotlib.pyplot as plt
import altair as alt
import openpyxl

downloads_path = str(Path.home() / "Downloads")
unlocked = False

@st.cache(allow_output_mutation=True)
def persistdata():
    return {}

@st.cache(allow_output_mutation=True)
def persistSheets():
    return []

@st.cache_data
def convert_df(df):
    # IMPORTANT: Cache the conversion to prevent computation on every rerun
    return df.to_csv(index=False).encode('utf-8')


def to_str(var):
    return str(list(np.reshape(np.asarray(var), (1, np.size(var)))[0]))[1:-1]

def _sum(arr):
    sum = 0

    for i in arr:
        sum = sum + i
 
    return(sum)


def returnTotalRegUnit(credit_units1, edited_df):
    dd_ar = []
    ttp = []

    for yy in range(len(edited_df)):
        ndf = edited_df.iloc[yy]
        cc = 0
        my_ntar = ndf.values
        dd = 0

        for i in my_ntar:
            try:
                if i == 'ABS':
                    dd + 0
                else:
                    dd += credit_units1[cc]
            except:
                pass
            cc += 1
        
        ttp.append(dd)
        dd = 0

    return ttp

def returnTotalPassed(credit_units1, edited_df):
    t_p_p = []

    for xx in range(len(edited_df)):
        jtt = _sum(credit_units1)
        ndfa = edited_df.iloc[xx]
        anpr = ndfa.values
        count = 0
        for i in anpr:
            try:
                if(int(i) >= 0 and int(i) <= 39):
                    jtt -= credit_units1[count]
            except:
                jtt -= credit_units1[count]
            count += 1

        t_p_p.append(jtt)
        jtt = 0

    return t_p_p

def returnTotalFailed(credit_units1, edited_df):
    t_u_f = []
    for yy in range(len(edited_df)):
        sss = 0
        ndfa1 = edited_df.iloc[yy]
        anpr1 = ndfa1.values
        count = 0
        for j in anpr1:
            try:
                if(int(j) <= 39):
                    sss += credit_units1[count]
            except:
                pass
            count += 1

        t_u_f.append(sss)
        sss = 0

    return t_u_f

def returnOutstandingCourses(credit_units1, edited_df):
    cols = edited_df.columns
    t_p_i_t = []

    for yy in range(len(edited_df)):
        sss = ""
        ndfa1 = edited_df.iloc[yy]
        anpr1 = ndfa1.values
        count = 0
        for idx, j in enumerate(anpr1):
            try:
                if(int(j) <= 39):
                    sss += cols[idx] + ','
            except:
                sss += cols[idx] + ','
            count += 1

        t_p_i_t.append(sss)
    return t_p_i_t


def returnTotalPoints(credit_units1, edited_df):
    t_p_i_t = []

    for yy in range(len(edited_df)):
        ndf = edited_df.iloc[yy]
        cc = 0
        my_ntar = ndf.values
        dd = 0

        for i in my_ntar:
            try:
                if(int(i) <= 100 and int(i) >= 70):
                    dd += credit_units1[cc] * 5
                elif(int(i) <= 69 and int(i) >= 60):
                    dd += credit_units1[cc] * 4
                elif(int(i) <= 59 and int(i) >= 50):
                    dd += credit_units1[cc] * 3
                elif(int(i) <= 49 and int(i) >= 45):
                    dd += credit_units1[cc] * 2
                elif(int(i) <= 44 and int(i) >= 40):
                    dd += credit_units1[cc] * 1
                elif(int(i) <= 39 and int(i) >= 0):
                    dd += credit_units1[cc] * 0
            except:
                dd += credit_units1[cc] * 0
            cc += 1

        
        t_p_i_t.append(dd)
        dd = 0
    return t_p_i_t


def returnOutstandingCoursesSummary(sheet_names, uploaded_file):
    out_cors = []
    for sn in range(len(sheet_names)-2):
        df = pd.read_excel(uploaded_file, sheet_name=sheet_names[sn])

        eac_out_cors = []
        for i in df['Outstanding Courses']:    
            eac_out_cors.append(i)
       
        out_cors.append(eac_out_cors)

    df1 = pd.DataFrame()

    for idx, value in enumerate(out_cors):
        df1[str(idx)] = value

    t_df = df1[df1.columns[0:]].apply(lambda x: ' '.join(x.dropna().astype(str)),axis=1)

    return t_df

def returnCoursesFailedSummary(sheet_names, uploaded_file):
    out_cors = []
    for sn in sheet_names[-2:]:
        df = pd.read_excel(uploaded_file, sheet_name=sn)

        eac_out_cors = []
        for i in df['Outstanding Courses']:    
            eac_out_cors.append(i)
       
        out_cors.append(eac_out_cors)

    df1 = pd.DataFrame()

    for idx, value in enumerate(out_cors):
        df1[str(idx)] = value

    t_df = df1[df1.columns[0:]].apply(lambda x: ' '.join(x.dropna().astype(str)),axis=1)

    return t_df

def returnTotalUnitsTakenSummary(sheet_names, uploaded_file):
    out_cors = []
    for sn in sheet_names[-2:]:
        df = pd.read_excel(uploaded_file, sheet_name=sn)

        eac_out_cors = []
        for i in df['Total Points Registered']:    
            eac_out_cors.append(i)
       
        out_cors.append(eac_out_cors)

    df1 = pd.DataFrame()

    for idx, value in enumerate(out_cors):
        df1[str(idx)] = value

    df1 = df1.astype('int64')
    t_df = df1.sum(axis=1, numeric_only= True)

    return t_df

def returnTotalUnitsPassedSummary(sheet_names, uploaded_file):
    out_cors = []
    for sn in sheet_names[-2:]:
        df = pd.read_excel(uploaded_file, sheet_name=sn)

        eac_out_cors = []
        for i in df['Total Points Passed']:    
            eac_out_cors.append(i)
       
        out_cors.append(eac_out_cors)

    df1 = pd.DataFrame()

    for idx, value in enumerate(out_cors):
        df1[str(idx)] = value

    df1 = df1.astype('int64')
    t_df = df1.sum(axis=1, numeric_only= True)

    return t_df

def returnSessionalGPASummary(sheet_names, uploaded_file):
    out_cors = []
    for sn in sheet_names[-2:]:
        df = pd.read_excel(uploaded_file, sheet_name=sn)

        eac_out_cors = []
        for i in df['Total Points Scored']:    
            eac_out_cors.append(i)
       
        out_cors.append(eac_out_cors)

    df1 = pd.DataFrame()

    for idx, value in enumerate(out_cors):
        df1[str(idx)] = value

    df1 = df1.astype('int64')
    t_df = df1.sum(axis=1, numeric_only= True)

    t_df = round(t_df, 3)
        

    ss = returnTotalUnitsTakenSummary(sheet_names, uploaded_file)

    new_arr = []

    for idx, val in enumerate(ss):
        aa = t_df[idx] / val
        aa = round(aa, 2)
        new_arr.append(aa)

    return new_arr

def returnCummulativeUnitsTakenSummary(sheet_names, uploaded_file):
    out_cors = []
    for sn in sheet_names:
        df = pd.read_excel(uploaded_file, sheet_name=sn)

        eac_out_cors = []
        for i in df['Total Points Registered']:    
            eac_out_cors.append(i)
       
        out_cors.append(eac_out_cors)

    df1 = pd.DataFrame()

    for idx, value in enumerate(out_cors):
        df1[str(idx)] = value

    df1 = df1.astype('int64')
    t_df = df1.sum(axis=1, numeric_only= True)


    return t_df

def returnCummulativeUnitsPassedSummary(sheet_names, uploaded_file):
    out_cors = []
    for sn in sheet_names:
        df = pd.read_excel(uploaded_file, sheet_name=sn)

        eac_out_cors = []
        for i in df['Total Points Passed']:    
            eac_out_cors.append(i)
       
        out_cors.append(eac_out_cors)

    df1 = pd.DataFrame()

    for idx, value in enumerate(out_cors):
        df1[str(idx)] = value

    df1 = df1.astype('int64')
    t_df = df1.sum(axis=1, numeric_only= True)

    return t_df

def returnCGPASummary(sheet_names, uploaded_file):
    out_cors = []
    for sn in sheet_names:
        df = pd.read_excel(uploaded_file, sheet_name=sn)

        eac_out_cors = []
        for i in df['Total Points Scored']:    
            eac_out_cors.append(i)
       
        out_cors.append(eac_out_cors)

    df1 = pd.DataFrame()

    for idx, value in enumerate(out_cors):
        df1[str(idx)] = value

    df1 = df1.astype('int64')
    t_df = df1.sum(axis=1, numeric_only= True)


    ss = returnCummulativeUnitsTakenSummary(sheet_names, uploaded_file)

    new_arr = []

    for idx, val in enumerate(ss):
        aa = t_df[idx] / val
        aa = round(aa, 2)
        new_arr.append(aa)

    return new_arr

def casual_calculator():

    st.title('CASUAL CALCULATOR')
    tcu = None

    credit_units = st.text_input('All Credit Accordingly Seperated By A Comma')
    grades = st.text_input('All Grades Accordingly Seperated By A Comma')

    if(st.button('Calculate')):
        st.info('Please Wait.....')

        cu = credit_units.split(',')
        sum = 0
        for i in cu:
            sum = sum + int(i)
            tcu = i
        
        st.text('Total Credit Unit --> ' + str(sum))

        grds = grades.split(',')
        grades_dict = {}

        total_a = []
        total_b = []
        total_c = []
        total_d = []
        total_e = []
        total_f = []

        for i in grds:
            if int(i) >= 70 and int(i) <= 100:
                total_a.append(i)
            elif int(i) >= 60 and int(i) <= 69:
                total_b.append(i)
            elif int(i) >= 50 and int(i) <= 59:
                total_c.append(i)
            elif int(i) >= 45 and int(i) <= 49:
                total_d.append(i)
            elif int(i) >= 40 and int(i) <= 44:
                total_e.append(i)
            elif int(i) >= 0 and int(i) <= 39:
                total_f.append(i)

        grades_dict.update({
            "A" : total_a,
            "B" : total_b,
            "C" : total_c,
            "D" : total_d,
            "E" : total_e,
            "F" : total_f
        })

        tps = 0

        aas = grades_dict.get('A')
        bbs = grades_dict.get('B')
        ccs = grades_dict.get('C')
        dds = grades_dict.get('D')
        ees = grades_dict.get('E')
        ffs = grades_dict.get('F')
        a = 0
        b = 0
        c = 0
        d = 0
        e = 0
        f = 0

        for i, value in enumerate(cu):
            if grds[int(i)] in (aas):
                a += int(value) * 5
            elif grds[int(i)] in bbs:
                b += int(value) * 4
            elif grds[int(i)] in ccs:
                c += int(value) * 3
            elif grds[int(i)] in dds:
                d += int(value) * 2
            elif grds[int(i)] in ees:
                e += int(value) * 1
            elif grds[int(i)] in ffs:
                f += int(value) * 0

        tps = a + b + c + d + e + f
        
        gpa = tps / sum

        st.text('Total Points Scored --> ' + str(tps))

        st.text('G.P --> ' + str(gpa))

def single_sheet_calculator():
    st.title('SINGLE SHEET CALCULATOR')

    if(st.button('View Sample Dataset Image')):
                st.info('Number Of Columns Dosen\'t Matter')
                st.image('sampledata.PNG')
    try:
        uploaded_file = st.file_uploader("Upload Dataset", type=['csv', 'xlsx'])
    except:
        pass

    if uploaded_file is not None:

        try:
            df = pd.read_csv(uploaded_file)
        except:
            wb = openpyxl.load_workbook(uploaded_file) 
            sheet_names = wb.sheetnames

            select_sheet = st.selectbox('Select Sheet:',sheet_names)
            df = pd.read_excel(uploaded_file, sheet_name=select_sheet)



        columns = st.multiselect("Columns:",df.columns)

        df1 = pd.DataFrame()

        for i in columns:
            df1[i] = df[i]
        
        # st.dataframe(edited_df)

        edited_df = st.experimental_data_editor(df1)
        
        ccr = st.text_input('Input All Credit Accordingly Seperated By A Comma')

        if(st.button('PROCEED')):

            cred_uni = ccr.split(',')
            cred_uni = [ int(x) for x in cred_uni]




            rtf = returnTotalFailed(cred_uni, edited_df)
            rtp = returnTotalPassed(cred_uni, edited_df)
            rtr = returnTotalRegUnit(cred_uni, edited_df)
            rtoc = returnOutstandingCourses(cred_uni, edited_df)
            rtsp = returnTotalPoints(cred_uni, edited_df)

            df['Total Points Registered'] = rtr
            df['Total Points Passed'] = rtp
            df['Total Points Failed'] = rtf
            df['Outstanding Courses'] = rtoc
            df['Total Points Scored'] = rtsp

            gp = []
            for i in range(len(df['Total Points Passed'])):
                tgp = df['Total Points Scored'][i] / df['Total Points Registered'][i]
                tgp = round(tgp, 3)
                gp.append(tgp)


            df['GPA'] = gp

            for i in edited_df.columns:
                df[i] = edited_df[i]


            st.dataframe(df.astype(str))

            ndf = df['GPA']

            res_ = []
            for i in df['GPA'].values:
                ss = ''
                if float(i) <= 4.5 and float(i) >= 5.0:
                    ss = ('First Class')
                elif float(i) <= 4.49 and float(i) >= 3.50:
                    ss = ('Second Class Upper')
                elif float(i) <= 3.49 and float(i) >= 2.40:
                    ss = ('Second Class Lower')
                elif float(i) <= 2.39 and float(i) >= 1.50:
                    ss = ('Third Class')
                elif float(i) <= 1.49 and float(i) >= 1.0:
                    ss = ('Pass')

                res_.append(ss)
            
            ndf['Grade'] = res_
            ndf['No Of Students'] = [i for i in range(len(ndf['Grade']))]

        

            # data1 = pd.DataFrame({
            #     'RANKING': ndf['Grade'],
            #     'No Of Students': ndf['No Of Students'],
            # })

            # chart = (alt.Chart(data1).mark_boxplot().encode(
            #     x=alt.X('RANKING', sort=None),
            #     y=alt.Y('No Of Students', sort=None)
            # ))

            # st.altair_chart(chart, use_container_width=True)

            
            aa = convert_df(df)

            file_name = 'mactechloop (' + uploaded_file.name[:-4].strip()+').csv'
        
    
            if(st.download_button(
            label="Download File",
            data=aa,
            file_name=file_name,
            mime='text/csv',
                )):
                st.success('Downloaded Successfully')

def multiple_sheets_calculator():
    st.title('MULTIPLE SHEET CALCULATOR')
    

    if(st.button('View Sample Dataset Image')):
        st.info('Number Of Columns Dosen\'t Matter')
        st.image('sampledata.PNG')
    try:
        uploaded_file = st.file_uploader("Upload Dataset", type=['xlsx'])
    except:
        pass

    dataframe_collection = persistdata()
    solved_sheets = persistSheets()
    

    if uploaded_file is not None:

        all_excel_data = pd.ExcelWriter('a1.xlsx', engine='xlsxwriter')
       
        widget_id = (id for id in range(1, 100_00))

        wb = openpyxl.load_workbook(uploaded_file) 
        sheet_names = wb.sheetnames

    
        # select_sheet = st.selectbox('Select Sheet:',sheet_names)
        # df = pd.read_excel(uploaded_file, sheet_name=select_sheet)

        tdf = pd.DataFrame()
        
         
        sht_nms = st.selectbox('Select Sheet', sheet_names)

        writer = pd.ExcelWriter('multiple.xlsx', engine='xlsxwriter')
    
        df = pd.read_excel(uploaded_file, sheet_name=sht_nms)

        columns = st.multiselect("Columns_Excel:",df.columns, key=next(widget_id))

        df1 = pd.DataFrame()
        for i in columns:
            df1[i] = df[i]

        edited_df = st.experimental_data_editor(df1, key=next(widget_id))
        
        ccr = st.text_input('Input All Credit Accordingly Seperated By A Comma', key=next(widget_id))

        if(st.button('PROCEED', key=next(widget_id))):

            cred_uni = ccr.split(',')
            cred_uni = [ int(x) for x in cred_uni]



            rtf = returnTotalFailed(cred_uni, edited_df)
            rtp = returnTotalPassed(cred_uni, edited_df)
            rtr = returnTotalRegUnit(cred_uni, edited_df)
            rotc = returnOutstandingCourses(cred_uni, edited_df)
            rtsp = returnTotalPoints(cred_uni, edited_df)

            df['Total Points Registered'] = rtr
            df['Total Points Passed'] = rtp
            df['Total Points Failed'] = rtf
            df['Outstanding Courses'] = rotc
            df['Total Points Scored'] = rtsp

            gp = []
            for i in range(len(df['Total Points Passed'])):
                tgp = df['Total Points Scored'][i] / df['Total Points Registered'][i]
                tgp = round(tgp, 3)
                gp.append(tgp)


            df['GPA'] = gp

            for i in edited_df.columns:
                df[i] = edited_df[i]

            st.dataframe(df)

            dic1 = {sht_nms : df}
            dataframe_collection.update(dic1)
            solved_sheets.append(sht_nms)
            
            
        st.write('Solved Sheets --> ' + str(solved_sheets))
        if(st.button('Download With Current Sheets', key=next(widget_id))):
            for key in dataframe_collection.keys():
                dataframe_collection[key].to_excel(all_excel_data, index=False, sheet_name=key)

            all_excel_data.save()

            with open('a1.xlsx', "rb") as template_file:
                template_byte = template_file.read()

                st.download_button(label="Click To Download",
                                    data=template_byte,
                                    file_name="template.xlsx",
                                    mime='application/octet-stream')
        
def summary():
    st.title('Data Summary')
    st.info('Upload The Dataset Downloaded From The Multiple Sheet Calculator To Get The Summary')
    try:
        uploaded_file = st.file_uploader("Upload Dataset", type=['csv', 'xlsx'])
    except:
        pass

    if uploaded_file is not None:

        try:
            df = pd.read_csv(uploaded_file)
        except:
            wb = openpyxl.load_workbook(uploaded_file) 
            sheet_names = wb.sheetnames

        new_df = pd.DataFrame()
        

        
        new_df['Outstanding Courses'] = returnOutstandingCoursesSummary(sheet_names, uploaded_file)
        new_df['Courses Failed'] = returnCoursesFailedSummary(sheet_names, uploaded_file)
        new_df['Total Units Taken'] = returnTotalUnitsTakenSummary(sheet_names, uploaded_file)
        new_df['Total Units Passed'] = returnTotalUnitsPassedSummary(sheet_names, uploaded_file)
        new_df['Sessional GPA'] = returnSessionalGPASummary(sheet_names, uploaded_file)
        new_df['Cummulative Units Taken'] = returnCummulativeUnitsTakenSummary(sheet_names, uploaded_file)
        new_df['Cummulative Units Passed'] = returnCummulativeUnitsPassedSummary(sheet_names, uploaded_file)
        new_df['CGPA'] = returnCGPASummary(sheet_names, uploaded_file)

        res_ = []

        for i in new_df['CGPA'].values:
            ss = ''
            if float(i) <= 4.5 and float(i) >= 5.0:
                ss = ('First Class')
            elif float(i) <= 4.49 and float(i) >= 3.50:
                ss = ('Second Class Upper')
            elif float(i) <= 3.49 and float(i) >= 2.40:
                ss = ('Second Class Lower')
            elif float(i) <= 2.39 and float(i) >= 1.50:
                ss = ('Third Class')
            elif float(i) <= 1.49 and float(i) >= 1.0:
                ss = ('Pass')

            res_.append(ss)

        new_df['Class Of Degree'] = res_

        rmk = []

        for idx, val in enumerate(new_df['Outstanding Courses']):
            if not len(val) == 0:
                rmk.append('FRNS')
            elif not len(new_df['Courses Failed'].iloc[idx]) == 0:
                rmk.append('FRNS')
            else:
                rmk.append('GRADUATING')

        new_df['Remark'] = rmk

        st.info('Select Other Columns To Make Up The Dataset (Example Name and Reg Number)')

        sht_nm = st.selectbox("Which Sheet Do You Want To Select From?", sheet_names)
        temp_df = pd.read_excel(uploaded_file, sheet_name=sht_nm)

        tr = list(temp_df.columns)
        tr.append('NO OTHER COLUMN NEEDED')

        name_temp_ = st.multiselect("Select Columns", tr)

        if(st.button('Continue')):
            st.dataframe(new_df)

            aa = convert_df(new_df)

            file_name = 'mactechloop (' + uploaded_file.name[:-4].strip()+').csv'

            if(name_temp_[0] == 'NO OTHER COLUMN NEEDED'):
                
                if(st.download_button(
                label="Download CSV File",
                data=aa,
                file_name=file_name,
                mime='text/csv',
                    )):
                    st.success('Downloaded Successfully')

            else:
                new_df[name_temp_] = temp_df[name_temp_]

                if(st.download_button(
                label="Download CSV File",
                data=aa,
                file_name=file_name,
                mime='text/csv',
                    )):
                    st.success('Downloaded Successfully')


def main():

    st.info('Built With ❤️ By MacTechLoop (mactechloop@gmail.com)')
    st.warning('Please Use The Nav Bar By The Top Right to Rerun Incase Of Any Glitches')

    activities=['Casual Calulator','Single Sheet Calculator','Multiple Sheet Calculator','Data Summary']
    option=st.sidebar.selectbox('Selection option:',activities)

    if option == activities[0]:
        casual_calculator()
    elif option == activities[1]:
        single_sheet_calculator()
    elif option == activities[2]:
        multiple_sheets_calculator()
    elif option == activities[3]:
        summary()


if __name__ == '__main__':
    main()
